import sys
import os
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLineEdit,
    QPushButton, QMessageBox, QLabel, QComboBox, QDateEdit,
    QTimeEdit, QGridLayout, QScrollArea, QGroupBox
)
from PySide6 import QtGui
from PySide6.QtCore import Qt, QDate, QTime
from PySide6.QtGui import QFont
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Функция для загрузки учетных номеров из Excel
def load_account_numbers(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    account_numbers = []
    
    # Получаем уже использованные номера из marshrutka.xlsx
    used_numbers = set()
    if os.path.exists('marshrutka.xlsx'):
        marshrutka_wb = load_workbook('marshrutka.xlsx')
        if "Records" in marshrutka_wb.sheetnames:
            marshrutka_sheet = marshrutka_wb["Records"]
            for row in marshrutka_sheet.iter_rows(min_row=2, values_only=True):
                if row[6]:  # Учетный номер в 7-м столбце (индекс 6)
                    used_numbers.add(row[6])
        marshrutka_wb.close()
    
    # Фильтруем номера из plavka.xlsx
    for row in sheet.iter_rows(min_row=2, values_only=True):
        account_number = row[1]  # Учетный номер во втором столбце
        if (account_number 
            and "/25" in str(account_number)  # Содержит "/25"
            and account_number not in used_numbers):  # Отсутствует в marshrutka.xlsx
            account_numbers.append(account_number)
    
    return sorted(account_numbers)  # Возвращаем отсортированный список

# Функция для сохранения данных в Excel
def save_to_excel(сборка_кластера_дата, сборка_кластера_специалист, сборка_кластера_количество,
                  контроль_сборки_кластера_дата_выставления, контроль_сборки_кластера_время_выставления,
                  контроль_сборки_кластера_специалист, учетный_номер, наименование_отливки, тип_эксперемента,
                  болгарка_дата, болгарка_специалист, термообработка_специалист,
                  дробеметная_обработка_специалист, зачищка_корона_специалист,
                  зачищка_лапа_специалист, зачищка_питатель_специалист, примечание):
    try:
        if os.path.exists('marshrutka.xlsx'):
            wb = load_workbook('marshrutka.xlsx')
            # Получаем лист Records или создаем его, если не существует
            if "Records" not in wb.sheetnames:
                ws = wb.create_sheet("Records")
                headers = ['Дата сборки', 'Специалист сборки', 'Количество', 
                          'Дата выставления', 'Время выставления', 'Специалист контроля',
                          'Учетный номер', 'Наименование отливки', 'Тип эксперимента',
                          'Дата болгарки', 'Специалист болгарки', 'Специалист термообработки',
                          'Специалист дробеметной обработки', 'Специалист зачистки короны',
                          'Специалист зачистки лапы', 'Специалист зачистки питателя', 'Примечание']
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)
            else:
                ws = wb["Records"]
            
            next_row = ws.max_row + 1
            
            data = [сборка_кластера_дата, сборка_кластера_специалист, сборка_кластера_количество,
                    контроль_сборки_кластера_дата_выставления, контроль_сборки_кластера_время_выставления,
                    контроль_сборки_кластера_специалист, учетный_номер, наименование_отливки, тип_эксперемента,
                    болгарка_дата, болгарка_специалист, термообработка_специалист,
                    дробеметная_обработка_специалист, зачищка_корона_специалист,
                    зачищка_лапа_специалист, зачищка_питатель_специалист, примечание]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=next_row, column=col)
                cell.value = value
                
                if col in [1, 4, 10]:  # Колонки с датами
                    cell.number_format = 'DD.MM.YYYY'
                elif col == 5:  # Колонка с временем
                    cell.number_format = 'HH:MM'
            
            wb.save('marshrutka.xlsx')
            wb.close()
        else:
            # Создаем новый файл если он не существует
            wb = Workbook()
            ws = wb.active
            ws.title = "Records"
            
            # Добавляем заголовки
            headers = ['Дата сборки', 'Специалист сборки', 'Количество', 
                      'Дата выставления', 'Время выставления', 'Специалист контроля',
                      'Учетный номер', 'Наименование отливки', 'Тип эксперимента',
                      'Дата болгарки', 'Специалист болгарки', 'Специалист термообработки',
                      'Специалист дробеметной обработки', 'Специалист зачистки короны',
                      'Специалист зачистки лапы', 'Специалист зачистки питателя', 'Примечание']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            
            # Добавляем данные
            data = [сборка_кластера_дата, сборка_кластера_специалист, сборка_кластера_количество,
                    контроль_сборки_кластера_дата_выставления, контроль_сборки_кластера_время_выставления,
                    контроль_сборки_кластера_специалист, учетный_номер, наименование_отливки, тип_эксперемента,
                    болгарка_дата, болгарка_специалист, термообработка_специалист,
                    дробеметная_обработка_специалист, зачищка_корона_специалист,
                    зачищка_лапа_специалист, зачищка_питатель_специалист, примечание]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=2, column=col)
                cell.value = value
                
                if col in [1, 4, 10]:  # Колонки с датами
                    cell.number_format = 'DD.MM.YYYY'
                elif col == 5:  # Колонка с временем
                    cell.number_format = 'HH:MM'
            
            wb.save('marshrutka.xlsx')
            wb.close()
    except Exception as e:
        raise Exception(f"Ошибка при сохранении в Excel: {str(e)}")


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Электронная маршрутная карта")
        # Установка фиксированного размера окна
        self.setFixedSize(1200, 800)
        
        # Основные цвета Aura
        self.BRAND_COLOR = "#0176D3"
        self.TEXT_COLOR = "#181818"
        self.BORDER_COLOR = "#DDDBDA"
        self.BG_COLOR = "#F3F3F3"
        self.SUCCESS_COLOR = "#45C65A"
        
        # Добавляем цвета для темной темы
        self.DARK_TEXT_COLOR = "#FFFFFF"
        self.DARK_BG_COLOR = "#1A1C1E"
        self.DARK_BORDER_COLOR = "#2D2D2D"
        
        # Флаг для отслеживания текущей темы
        self.is_dark_theme = False
        
        # Создаем переключатель темы
        self.theme_toggle = QPushButton(self)
        self.theme_toggle.setFixedSize(60, 30)
        self.theme_toggle.setCheckable(True)
        self.theme_toggle.clicked.connect(self.toggle_theme)
        self.theme_toggle.setStyleSheet("""
            QPushButton {
                background-color: #DDDBDA;
                border: none;
                border-radius: 15px;
                padding: 2px;
                text-align: left;
            }
            QPushButton:checked {
                background-color: #0176D3;
                text-align: right;
            }
            QPushButton::indicator {
                width: 26px;
                height: 26px;
                background-color: white;
                border-radius: 13px;
                margin: 2px;
            }
            QPushButton:checked::indicator {
                background-color: white;
            }
        """)
        
        # Установка основного стиля окна
        self.setStyleSheet(f"""
            QWidget {{
                background-color: {self.BG_COLOR};
                font-family: 'Segoe UI', Arial;
                font-size: 13px;
                color: {self.TEXT_COLOR};
            }}
            QLabel {{
                padding: 4px 0;
                font-size: 13px;
            }}
            QLineEdit, QComboBox, QDateEdit {{
                padding: 4px 8px;
                border: 1px solid {self.BORDER_COLOR};
                border-radius: 4px;
                background-color: white;
                min-height: 24px;
            }}
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{
                border: 2px solid {self.BRAND_COLOR};
                outline: none;
            }}
            QComboBox::drop-down {{
                border: none;
                width: 24px;
            }}
            QPushButton {{
                background-color: {self.BRAND_COLOR};
                color: white;
                padding: 6px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                min-height: 24px;
            }}
            QPushButton:hover {{
                background-color: #014486;
            }}
            QPushButton:pressed {{
                background-color: #032D60;
            }}
        """)

        # Списки специалистов
        scleyks = ["Буцик", "Минакова", "Ротарь", "Чернова", "Чупахина"]
        controlers = ["Елхова", "Шестункина", "Романцева"]
        bolgar = [
            "Ахмаджонов", "Отаназаров", "Косимов", "Косимов-2", "Туичев",
            "Машрапов", "Эргашев", "Самиев", "Исмаилов"
        ]
        termob = ["Эгамов", "Аюбов"]
        drobem = ["Эгамов", "Аюбов"]
        zachistka = ["Абдуллаев", "Бурхонов", "Матесаев", "Мещерякова",
            "Самиев", "Леонтьева"]
        
        # Инициализация всех полей формы
        self.сборка_кластера_дата = QDateEdit(self)
        self.сборка_кластера_дата.setDisplayFormat("dd.MM.yyyy")
        self.сборка_кластера_дата.setCalendarPopup(True)
        self.сборка_кластера_дата.setDate(QDate.currentDate())

        self.сборка_кластера_специалист = QComboBox(self)
        self.сборка_кластера_специалист.addItems(scleyks)
        self.сборка_кластера_специалист.setCurrentIndex(-1)
        self.сборка_кластера_специалист.setPlaceholderText("Специалист по сборке кластера")

        self.сборка_кластера_количество = QLineEdit(self)
        self.сборка_кластера_количество.setPlaceholderText("Количество кластера")

        self.контроль_сборки_кластера_дата_выставления = QDateEdit(self)
        self.контроль_сборки_кластера_дата_выставления.setDisplayFormat("dd.MM.yyyy")
        self.контроль_сборки_кластера_дата_выставления.setCalendarPopup(True)
        self.контроль_сборки_кластера_дата_выставления.setDate(QDate.currentDate())

        self.контроль_сборки_кластера_время_выставления = QTimeEdit(self)
        self.контроль_сборки_кластера_время_выставления.setDisplayFormat("HH:mm")

        self.контроль_сборки_кластера_специалист = QComboBox(self)
        self.контроль_сборки_кластера_специалист.addItems(controlers)
        self.контроль_сборки_кластера_специалист.setCurrentIndex(-1)
        self.контроль_сборки_кластера_специалист.setPlaceholderText("Специалист по контролю кластера")

        self.учетный_номер = QComboBox(self)
        self.учетный_номер.addItems(load_account_numbers('plavka.xlsx'))
        self.учетный_номер.currentTextChanged.connect(self.update_experiment_details)

        self.наименование_отливки = QLineEdit(self)
        self.наименование_отливки.setPlaceholderText("Наименование отливки")

        self.тип_эксперемента = QLineEdit(self)
        self.тип_эксперемента.setPlaceholderText("Тип эксперимента")

        self.болгарка_дата = QDateEdit(self)
        self.болгарка_дата.setDisplayFormat("dd.MM.yyyy")
        self.болгарка_дата.setCalendarPopup(True)
        self.болгарка_дата.setDate(QDate.currentDate())

        self.болгарка_специалист = QComboBox(self)
        self.болгарка_специалист.addItems(bolgar)
        self.болгарка_специалист.setCurrentIndex(-1)
        self.болгарка_специалист.setPlaceholderText("Специалист по резке")

        self.термообработка_специалист = QComboBox(self)
        self.термообработка_специалист.addItems(termob)
        self.термообработка_специалист.setCurrentIndex(-1)
        self.термообработка_специалист.setPlaceholderText("Специалист по термообработке")

        self.дробеметная_обработка_специалист = QComboBox(self)
        self.дробеметная_обработка_специалист.addItems(drobem)
        self.дробеметная_обработка_специалист.setCurrentIndex(-1)
        self.дробеметная_обработка_специалист.setPlaceholderText("Специалист по дробеметной обработке")

        self.зачистка_корона_специалист = QComboBox(self)
        self.зачистка_корона_специалист.addItems(zachistka)
        self.зачистка_корона_специалист.setCurrentIndex(-1)
        self.зачистка_корона_специалист.setPlaceholderText("Специалист по зачистке короны")

        self.зачистка_лапа_специалист = QComboBox(self)
        self.зачистка_лапа_специалист.addItems(zachistka)
        self.зачистка_лапа_специалист.setCurrentIndex(-1)
        self.зачистка_лапа_специалист.setPlaceholderText("Специалист по зачистке лапы")

        self.зачистка_питатель_специалист = QComboBox(self)
        self.зачистка_питатель_специалист.addItems(zachistka)
        self.зачистка_питатель_специалист.setCurrentIndex(-1)
        self.зачистка_питатель_специалист.setPlaceholderText("Специалист по зачистке питателя")

        self.примечание = QLineEdit(self)
        self.примечание.setPlaceholderText("Примечание")

        # Создаем основной layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(16, 16, 16, 16)

        # Добавляем горизонтальный layout для заголовка и переключателя
        header_layout = QGridLayout()
        
        # Заголовок
        title = QLabel("МАРШРУТНАЯ КАРТА")
        title.setStyleSheet(f"""
            font-size: 20px;
            font-weight: bold;
            color: {self.BRAND_COLOR};
            padding: 0 0 8px 0;
            border-bottom: 2px solid {self.BORDER_COLOR};
            margin-bottom: 8px;
        """)
        header_layout.addWidget(title, 0, 0)
        header_layout.addWidget(self.theme_toggle, 0, 1, Qt.AlignRight)
        main_layout.addLayout(header_layout)

        # Создаем область содержимого
        content = QWidget()
        grid_layout = QGridLayout(content)
        grid_layout.setSpacing(12)  # Увеличиваем расстояние между элементами
        
        # Функция для создания группы полей
        def create_group(title):
            group = QGroupBox(title)
            group.setStyleSheet(f"""
                QGroupBox {{
                    border: 1px solid {self.BORDER_COLOR};
                    border-radius: 6px;
                    margin-top: 12px;
                    padding: 12px;
                    background-color: white;
                }}
                QGroupBox::title {{
                    color: {self.BRAND_COLOR};
                    padding: 0 8px;
                    background-color: white;
                    font-weight: bold;
                }}
            """)
            return group

        # Левая колонка
        # Группа "Сборка кластера"
        сборка_group = create_group("Сборка кластера")
        сборка_layout = QGridLayout()
        сборка_layout.addWidget(QLabel("Дата:"), 0, 0)
        сборка_layout.addWidget(self.сборка_кластера_дата, 0, 1)
        сборка_layout.addWidget(QLabel("Специалист:"), 1, 0)
        сборка_layout.addWidget(self.сборка_кластера_специалист, 1, 1)
        сборка_layout.addWidget(QLabel("Количество:"), 2, 0)
        сборка_layout.addWidget(self.сборка_кластера_количество, 2, 1)
        сборка_group.setLayout(сборка_layout)
        grid_layout.addWidget(сборка_group, 0, 0)

        # Группа "Контроль сборки"
        контроль_group = create_group("Контроль сборки")
        контроль_layout = QGridLayout()
        контроль_layout.addWidget(QLabel("Дата:"), 0, 0)
        контроль_layout.addWidget(self.контроль_сборки_кластера_дата_выставления, 0, 1)
        контроль_layout.addWidget(QLabel("Время:"), 1, 0)
        контроль_layout.addWidget(self.контроль_сборки_кластера_время_выставления, 1, 1)
        контроль_layout.addWidget(QLabel("Специалист:"), 2, 0)
        контроль_layout.addWidget(self.контроль_сборки_кластера_специалист, 2, 1)
        контроль_group.setLayout(контроль_layout)
        grid_layout.addWidget(контроль_group, 1, 0)

        # Группа "Информация об отливке"
        отливка_group = create_group("Информация об отливке")
        отливка_layout = QGridLayout()
        отливка_layout.addWidget(QLabel("Учетный номер:"), 0, 0)
        отливка_layout.addWidget(self.учетный_номер, 0, 1)
        отливка_layout.addWidget(QLabel("Наименование:"), 1, 0)
        отливка_layout.addWidget(self.наименование_отливки, 1, 1)
        отливка_layout.addWidget(QLabel("Тип эксперимента:"), 2, 0)
        отливка_layout.addWidget(self.тип_эксперемента, 2, 1)
        отливка_group.setLayout(отливка_layout)
        grid_layout.addWidget(отливка_group, 2, 0)

        # Правая колонка
        # Группа "Обработка"
        обработка_group = create_group("Обработка")
        обработка_layout = QGridLayout()
        обработка_layout.addWidget(QLabel("Дата болгарки:"), 0, 0)
        обработка_layout.addWidget(self.болгарка_дата, 0, 1)
        обработка_layout.addWidget(QLabel("Специалист болгарки:"), 1, 0)
        обработка_layout.addWidget(self.болгарка_специалист, 1, 1)
        обработка_layout.addWidget(QLabel("Специалист термообработки:"), 2, 0)
        обработка_layout.addWidget(self.термообработка_специалист, 2, 1)
        обработка_layout.addWidget(QLabel("Специалист дробеметки:"), 3, 0)
        обработка_layout.addWidget(self.дробеметная_обработка_специалист, 3, 1)
        обработка_group.setLayout(обработка_layout)
        grid_layout.addWidget(обработка_group, 0, 1)

        # Группа "Зачистка"
        зачистка_group = create_group("Зачистка")
        зачистка_layout = QGridLayout()
        зачистка_layout.addWidget(QLabel("Специалист зачистки короны:"), 0, 0)
        зачистка_layout.addWidget(self.зачистка_корона_специалист, 0, 1)
        зачистка_layout.addWidget(QLabel("Специалист зачистки лапы:"), 1, 0)
        зачистка_layout.addWidget(self.зачистка_лапа_специалист, 1, 1)
        зачистка_layout.addWidget(QLabel("Специалист зачистки питателя:"), 2, 0)
        зачистка_layout.addWidget(self.зачистка_питатель_специалист, 2, 1)
        зачистка_group.setLayout(зачистка_layout)
        grid_layout.addWidget(зачистка_group, 1, 1, 2, 1)

        # Группа "Дополнительно" (внизу на всю ширину)
        доп_group = create_group("Дополнительно")
        доп_layout = QGridLayout()
        доп_layout.addWidget(QLabel("Примечание:"), 0, 0)
        self.примечание.setFixedHeight(60)
        доп_layout.addWidget(self.примечание, 0, 1)
        доп_group.setLayout(доп_layout)
        grid_layout.addWidget(доп_group, 3, 0, 1, 2)

        # Добавляем содержимое в основной layout
        main_layout.addWidget(content)

        # Кнопка сохранения
        self.save_button = QPushButton("Сохранить", self)
        self.save_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.SUCCESS_COLOR};
                color: white;
                padding: 8px 24px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 14px;
                min-height: 32px;
            }}
            QPushButton:hover {{
                background-color: #2E844A;
            }}
            QPushButton:pressed {{
                background-color: #194E31;
            }}
        """)
        self.save_button.clicked.connect(self.save_data)
        main_layout.addWidget(self.save_button)

        self.setLayout(main_layout)

    def update_experiment_details(self):
        """Обновляет поля на основании выбранного учетного номера"""
        try:
            if os.path.exists('plavka.xlsx'):
                wb = load_workbook('plavka.xlsx')
                sheet = wb.active
                account_number = self.учетный_номер.currentText()
                
                # Ищем строку с выбранным учетным номером
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[1] == account_number:  # Учетный_номер во втором столбце
                        # Наименование отливки в 11-м столбце (индекс 10)
                        if row[10]:  # Наименование_отливки
                            self.наименование_отливки.setText(str(row[10]))
                        
                        # Тип эксперимента в 12-м столбце (индекс 11)
                        if row[11]:  # Тип_эксперемента
                            self.тип_эксперемента.setText(str(row[11]))
                        
                        # Дата плавки в 3-м столбце (индекс 2)
                        плавка_дата = row[2]  # Плавка_дата
                        
                        if isinstance(плавка_дата, datetime):
                            # Если дата в формате datetime
                            qdate = QDate(плавка_дата.year, плавка_дата.month, плавка_дата.day)
                            self.болгарка_дата.setDate(qdate)
                        elif isinstance(плавка_дата, str):
                            # Если дата в строковом формате DD.MM.YYYY
                            try:
                                day, month, year = map(int, плавка_дата.split('.'))
                                qdate = QDate(year, month, day)
                                if qdate.isValid():
                                    self.болгарка_дата.setDate(qdate)
                            except Exception as e:
                                print(f"Ошибка при преобразовании даты: {e}")
                        break
                
                wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при обновлении данных: {str(e)}")

    def validate_time(self, time_str):
        """Проверка корректности ввода времени в формате ЧЧ:ММ"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            if 0 <= hours < 24 and 0 <= minutes < 60:
                return True
        except ValueError:
            return False
        return False

    def save_data(self):
        # Проверяем обязательные поля
        required_fields = [
            (self.учетный_номер, "Учетный номер"),
            (self.наименование_отливки, "Наименование отливки"),
            (self.тип_эксперемента, "Тип эксперимента"),
            (self.сборка_кластера_специалист, "Специалист сборки"),
            (self.контроль_сборки_кластера_специалист, "Специалист контроля")
        ]
        
        empty_fields = []
        for field, name in required_fields:
            if isinstance(field, QComboBox):
                if field.currentText() == "":
                    empty_fields.append(name)
            elif isinstance(field, QLineEdit):
                if field.text().strip() == "":
                    empty_fields.append(name)
        
        if empty_fields:
            QMessageBox.warning(
                self,
                "Не все поля заполнены",
                "Пожалуйста, заполните следующие обязательные поля:\n• " + "\n• ".join(empty_fields)
            )
            return

        сборка_кластера_дата = self.сборка_кластера_дата.date().toString("dd.MM.yyyy")
        сборка_кластера_специалист = self.сборка_кластера_специалист.currentText().strip()
        сборка_кластера_количество = self.сборка_кластера_количество.text().strip()
        контроль_сборки_кластера_дата_выставления = self.контроль_сборки_кластера_дата_выставления.date().toString("dd.MM.yyyy")
        контроль_сборки_кластера_время_выставления = self.контроль_сборки_кластера_время_выставления.time().toString("HH:mm")
        контроль_сборки_кластера_специалист = self.контроль_сборки_кластера_специалист.currentText().strip()
        учетный_номер = self.учетный_номер.currentText().strip()
        болгарка_дата = self.болгарка_дата.date().toString("dd.MM.yyyy")
        болгарка_специалист = self.болгарка_специалист.currentText().strip()
        термообработка_специалист = self.термообработка_специалист.currentText().strip()
        дробеметная_обработка_специалист = self.дробеметная_обработка_специалист.currentText().strip()
        зачищка_корона_специалист = self.зачистка_корона_специалист.currentText().strip()
        зачищка_лапа_специалист = self.зачистка_лапа_специалист.currentText().strip()
        зачищка_питатель_специалист = self.зачистка_питатель_специалист.currentText().strip()
        примечание = self.примечание.text().strip()

        # Валидация времени
        if not self.validate_time(контроль_сборки_кластера_время_выставления):
            QMessageBox.warning(self, "Ошибка", "Некорректный ввод времени. Используйте формат ЧЧ:ММ.")
            return

        # Получение наименования отливки и типа эксперимента
        наименование_отливки = self.наименование_отливки.text().strip()
        тип_эксперемента = self.тип_эксперемента.text().strip()

        save_to_excel(сборка_кластера_дата, сборка_кластера_специалист, сборка_кластера_количество,
                       контроль_сборки_кластера_дата_выставления, контроль_сборки_кластера_время_выставления,
                       контроль_сборки_кластера_специалист, учетный_номер, наименование_отливки, тип_эксперемента,
                       болгарка_дата, болгарка_специалист, термообработка_специалист,
                       дробеметная_обработка_специалист, зачищка_корона_специалист,
                       зачищка_лапа_специалист, зачищка_питатель_специалист, примечание)

        QMessageBox.information(self, "Успех", "Данные сохранены в Excel!")

        # Обновляем список учетных номеров
        current_account_numbers = load_account_numbers('plavka.xlsx')
        self.учетный_номер.clear()
        self.учетный_номер.addItems(current_account_numbers)

        # Очистка полей ввода
        self.clear_fields()

    def clear_fields(self):
        """Очищает все поля формы"""
        # Очистка текстовых полей
        self.наименование_отливки.clear()
        self.тип_эксперемента.clear()
        self.примечание.clear()
        
        # Очистка комбобоксов
        self.сборка_кластера_специалист.setCurrentIndex(-1)
        self.контроль_сборки_кластера_специалист.setCurrentIndex(-1)
        self.болгарка_специалист.setCurrentIndex(-1)
        self.термообработка_специалист.setCurrentIndex(-1)
        self.дробеметная_обработка_специалист.setCurrentIndex(-1)
        self.зачистка_корона_специалист.setCurrentIndex(-1)
        self.зачистка_лапа_специалист.setCurrentIndex(-1)
        self.зачистка_питатель_специалист.setCurrentIndex(-1)
        
        # Установка пустых дат (текущая дата)
        empty_date = QDate.currentDate()
        self.сборка_кластера_дата.setDate(empty_date)
        self.контроль_сборки_кластера_дата_выставления.setDate(empty_date)
        self.болгарка_дата.setDate(empty_date)
        
        # Установка пустого времени (00:00)
        self.контроль_сборки_кластера_время_выставления.setTime(QTime(0, 0))
        
        # Очистка количества
        self.сборка_кластера_количество.clear()

    def toggle_theme(self):
        self.is_dark_theme = self.theme_toggle.isChecked()
        
        # Обновляем цвета в зависимости от темы
        text_color = self.DARK_TEXT_COLOR if self.is_dark_theme else self.TEXT_COLOR
        bg_color = self.DARK_BG_COLOR if self.is_dark_theme else self.BG_COLOR
        border_color = self.DARK_BORDER_COLOR if self.is_dark_theme else self.BORDER_COLOR
        
        # Обновляем стили для всего приложения
        self.setStyleSheet(f"""
            QWidget {{
                background-color: {bg_color};
                font-family: 'Segoe UI', Arial;
                font-size: 13px;
                color: {text_color};
            }}
            QLabel {{
                padding: 4px 0;
                font-size: 13px;
            }}
            QLineEdit, QComboBox, QDateEdit {{
                padding: 4px 8px;
                border: 1px solid {border_color};
                border-radius: 4px;
                background-color: {bg_color};
                color: {text_color};
                min-height: 24px;
            }}
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{
                border: 2px solid {self.BRAND_COLOR};
                outline: none;
            }}
            QComboBox::drop-down {{
                border: none;
                width: 24px;
            }}
            QPushButton {{
                background-color: {self.BRAND_COLOR};
                color: white;
                padding: 6px 16px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                min-height: 24px;
            }}
            QPushButton:hover {{
                background-color: #014486;
            }}
            QPushButton:pressed {{
                background-color: #032D60;
            }}
            QGroupBox {{
                border: 1px solid {border_color};
                border-radius: 6px;
                margin-top: 12px;
                padding: 12px;
                background-color: {bg_color};
            }}
            QGroupBox::title {{
                color: {self.BRAND_COLOR};
                padding: 0 8px;
                background-color: {bg_color};
                font-weight: bold;
            }}
        """)
        
        # Обновляем стиль заголовка
        for widget in self.findChildren(QLabel):
            if widget.text() == "МАРШРУТНАЯ КАРТА":
                widget.setStyleSheet(f"""
                    font-size: 20px;
                    font-weight: bold;
                    color: {self.BRAND_COLOR};
                    padding: 0 0 8px 0;
                    border-bottom: 2px solid {border_color};
                    margin-bottom: 8px;
                """)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
